from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from itertools import product
from statistics import pstdev
from typing import List, Optional
import copy

WEEKS = 35
ROLE_NAMES = ["Retailer", "Wholesaler", "Distributor", "Factory"]
BASE_DEMAND_35 = [100] * WEEKS
LOW_DEMAND_35 = [
    64, 96, 76, 88, 72, 104, 68, 92, 80, 84,
    60, 100, 78, 82, 70, 90, 74, 86, 56, 112,
    77, 83, 69, 91, 75, 85, 66, 94, 78, 82,
    71, 89, 74, 86, 48,
]
HIGH_DEMAND_35 = [
    91, 139, 109, 127, 103, 151, 97, 133, 115, 121,
    85, 145, 113, 117, 101, 129, 105, 125, 79, 163,
    110, 120, 98, 132, 108, 122, 93, 137, 111, 119,
    102, 128, 107, 123, 67,
]


def _simulate_cost_components_for_targets(targets: tuple, demand_series: list) -> tuple:
    inv_cost = 0.25
    back_cost = 1.0
    expected_demand = 100

    state = {
        role: {
            "inv": 100,
            "backlog": 0,
            "inbound_pipe": [50, 50],      # ship/prod delay
            "order_pipe": [0, 0],          # order delay
            "total_cost": 0.0,
        }
        for role in ROLE_NAMES
    }

    for demand in demand_series:
        # A) Receive inbound goods
        received = {}
        for role in ROLE_NAMES:
            arriving = state[role]["inbound_pipe"].pop(0)
            state[role]["inv"] += arriving
            state[role]["inbound_pipe"].append(0)
            received[role] = arriving

        # B) Incoming demand/order
        incoming = {
            "Retailer": demand,
            "Wholesaler": state["Retailer"]["order_pipe"].pop(0),
            "Distributor": state["Wholesaler"]["order_pipe"].pop(0),
            "Factory": state["Distributor"]["order_pipe"].pop(0),
        }
        for role in ROLE_NAMES:
            state[role]["order_pipe"].append(0)

        # C) Fulfill backlog + incoming
        shipped = {}
        for role in ROLE_NAMES:
            required = state[role]["backlog"] + incoming[role]
            ship = min(state[role]["inv"], required)
            state[role]["inv"] -= ship
            state[role]["backlog"] = required - ship
            shipped[role] = ship

        # D) Costs
        for role in ROLE_NAMES:
            week_c = state[role]["inv"] * inv_cost + state[role]["backlog"] * back_cost
            state[role]["total_cost"] += week_c

        # E) Place order (base-stock rule)
        orders = {}
        for i, role in enumerate(ROLE_NAMES):
            target_level = expected_demand * targets[i]
            inventory_position = state[role]["inv"] + sum(state[role]["inbound_pipe"]) - state[role]["backlog"]
            order_qty = max(0, int(target_level - inventory_position))
            orders[role] = order_qty
            state[role]["order_pipe"][-1] += order_qty

        # G) Shipments scheduled to downstream
        state["Retailer"]["inbound_pipe"][-1] += shipped["Wholesaler"]
        state["Wholesaler"]["inbound_pipe"][-1] += shipped["Distributor"]
        state["Distributor"]["inbound_pipe"][-1] += shipped["Factory"]
        state["Factory"]["inbound_pipe"][-1] += orders["Factory"]  # production completion

    total = sum(state[role]["total_cost"] for role in ROLE_NAMES)
    inv_total = sum(state[role]["inv"] * 0 for role in ROLE_NAMES)  # placeholder to keep signature simple
    # Recompute components from cumulative totals is not directly tracked above, so track in loop instead.
    # To avoid intrusive refactor, run a second pass of accounting with explicit accumulators.
    inv_acc = 0.0
    back_acc = 0.0
    state2 = {
        role: {"inv": 100, "backlog": 0, "inbound_pipe": [50, 50], "order_pipe": [0, 0]}
        for role in ROLE_NAMES
    }
    for demand in demand_series:
        for role in ROLE_NAMES:
            arriving = state2[role]["inbound_pipe"].pop(0)
            state2[role]["inv"] += arriving
            state2[role]["inbound_pipe"].append(0)
        incoming = {
            "Retailer": demand,
            "Wholesaler": state2["Retailer"]["order_pipe"].pop(0),
            "Distributor": state2["Wholesaler"]["order_pipe"].pop(0),
            "Factory": state2["Distributor"]["order_pipe"].pop(0),
        }
        for role in ROLE_NAMES:
            state2[role]["order_pipe"].append(0)
        shipped = {}
        for role in ROLE_NAMES:
            required = state2[role]["backlog"] + incoming[role]
            ship = min(state2[role]["inv"], required)
            state2[role]["inv"] -= ship
            state2[role]["backlog"] = required - ship
            shipped[role] = ship
            inv_acc += state2[role]["inv"] * inv_cost
            back_acc += state2[role]["backlog"] * back_cost
        orders = {}
        for i, role in enumerate(ROLE_NAMES):
            target_level = expected_demand * targets[i]
            inventory_position = state2[role]["inv"] + sum(state2[role]["inbound_pipe"]) - state2[role]["backlog"]
            order_qty = max(0, int(target_level - inventory_position))
            orders[role] = order_qty
            state2[role]["order_pipe"][-1] += order_qty
        state2["Retailer"]["inbound_pipe"][-1] += shipped["Wholesaler"]
        state2["Wholesaler"]["inbound_pipe"][-1] += shipped["Distributor"]
        state2["Distributor"]["inbound_pipe"][-1] += shipped["Factory"]
        state2["Factory"]["inbound_pipe"][-1] += orders["Factory"]
    return total, inv_acc, back_acc


def _optimize_targets(
    demand_series: list,
    min_tw: int,
    max_tw: int,
    prefer_distribution: bool = True,
    objective: str = "total",
) -> tuple:
    candidates = []
    best_score = float("inf")
    for t in product(range(min_tw, max_tw + 1), repeat=4):
        total_c, inv_c, back_c = _simulate_cost_components_for_targets(t, demand_series)
        if objective == "backlog":
            score = back_c
        elif objective == "inventory":
            score = inv_c
        else:
            score = total_c
        if score < best_score:
            best_score = score
        candidates.append((t, score, total_c, inv_c, back_c))

    if not prefer_distribution:
        best = min(candidates, key=lambda x: (x[1], x[2], x[0]))
        return best[0]

    # Keep near-optimal policies, then choose most distributed target weeks.
    near = [x for x in candidates if x[1] <= best_score * 1.01]
    if not near:
        near = candidates
    near.sort(key=lambda x: (-len(set(x[0])), -pstdev(x[0]), x[1], x[2], x[0]))
    return near[0][0]


def _init_state() -> dict:
    return {
        role: {
            "inv": 100,
            "backlog": 0,
            "inbound_pipe": [50, 50],
            "order_pipe": [0, 0],
            "cum_cost": 0.0,
        }
        for role in ROLE_NAMES
    }


def _step_state(state: dict, demand: int, targets: tuple, collect: bool = False):
    inv_cost = 0.25
    back_cost = 1.0
    expected_demand = 100

    # A) Receive inbound goods
    received = {}
    beg_inv = {}
    beg_back = {}
    for role in ROLE_NAMES:
        beg_inv[role] = state[role]["inv"]
        beg_back[role] = state[role]["backlog"]
        arriving = state[role]["inbound_pipe"].pop(0)
        state[role]["inv"] += arriving
        state[role]["inbound_pipe"].append(0)
        received[role] = arriving

    # B) Incoming demand/order
    incoming = {
        "Retailer": demand,
        "Wholesaler": state["Retailer"]["order_pipe"].pop(0),
        "Distributor": state["Wholesaler"]["order_pipe"].pop(0),
        "Factory": state["Distributor"]["order_pipe"].pop(0),
    }
    for role in ROLE_NAMES:
        state[role]["order_pipe"].append(0)

    # C) Fulfill backlog + incoming
    shipped = {}
    week_cost = {}
    inv_part = {}
    back_part = {}
    total_req = {}
    for role in ROLE_NAMES:
        required = state[role]["backlog"] + incoming[role]
        ship = min(state[role]["inv"], required)
        state[role]["inv"] -= ship
        state[role]["backlog"] = required - ship
        shipped[role] = ship
        total_req[role] = required
        inv_c = state[role]["inv"] * inv_cost
        back_c = state[role]["backlog"] * back_cost
        week_c = inv_c + back_c
        inv_part[role] = inv_c
        back_part[role] = back_c
        week_cost[role] = week_c
        state[role]["cum_cost"] += week_c

    # E) Place order (base-stock)
    orders = {}
    for i, role in enumerate(ROLE_NAMES):
        target_level = expected_demand * targets[i]
        inventory_position = state[role]["inv"] + sum(state[role]["inbound_pipe"]) - state[role]["backlog"]
        order_qty = max(0, int(target_level - inventory_position))
        orders[role] = order_qty
        state[role]["order_pipe"][-1] += order_qty

    # G) Shipments scheduled downstream
    state["Retailer"]["inbound_pipe"][-1] += shipped["Wholesaler"]
    state["Wholesaler"]["inbound_pipe"][-1] += shipped["Distributor"]
    state["Distributor"]["inbound_pipe"][-1] += shipped["Factory"]
    state["Factory"]["inbound_pipe"][-1] += orders["Factory"]

    if not collect:
        return None

    rows = {}
    for role in ROLE_NAMES:
        rows[role] = {
            "Incoming": incoming[role],
            "Recv": received[role],
            "BegInv": beg_inv[role],
            "BegBacklog": beg_back[role],
            "TotalReq": total_req[role],
            "Ship": shipped[role],
            "EndInv": state[role]["inv"],
            "EndBacklog": state[role]["backlog"],
            "InvCost": inv_part[role],
            "BackCost": back_part[role],
            "WeekCost": week_cost[role],
            "Cumulative": state[role]["cum_cost"],
            "OrderPlaced": orders[role],
        }
    return rows


def _horizon_cost_from_state(base_state: dict, demand_seq: list, targets: tuple, objective: str = "total") -> float:
    state = copy.deepcopy(base_state)
    total_cost = 0.0
    inv_cost = 0.0
    back_cost = 0.0
    for d in demand_seq:
        rows = _step_state(state, d, targets, collect=True)
        for role in ROLE_NAMES:
            total_cost += rows[role]["WeekCost"]
            inv_cost += rows[role]["InvCost"]
            back_cost += rows[role]["BackCost"]
    if objective == "backlog":
        return back_cost
    if objective == "inventory":
        return inv_cost
    return total_cost


def _best_targets_from_state(
    base_state: dict,
    demand_seq: list,
    min_tw: int,
    max_tw: int,
    objective: str = "total",
    prefer_distribution: bool = False,
) -> tuple:
    best_t = None
    best_cost = float("inf")
    candidates = []
    for t in product(range(min_tw, max_tw + 1), repeat=4):
        c = _horizon_cost_from_state(base_state, demand_seq, t, objective=objective)
        candidates.append((t, c))
        if c < best_cost:
            best_cost = c
            best_t = t
    if not prefer_distribution:
        return best_t
    near = [x for x in candidates if x[1] <= best_cost * 1.01]
    if not near:
        near = candidates
    near.sort(key=lambda x: (-len(set(x[0])), -pstdev(x[0]), x[1], x[0]))
    return near[0][0]


def _rolling_plan_rows(
    demand_series: list,
    min_tw: int = 1,
    max_tw: int = 6,
    horizon: int = 6,
    objective: str = "total",
    prefer_distribution: bool = False,
):
    state = _init_state()
    out_rows = []
    for w, demand in enumerate(demand_series, start=1):
        look = demand_series[w - 1: min(len(demand_series), w - 1 + horizon)]
        best_t = _best_targets_from_state(
            state,
            look,
            min_tw=min_tw,
            max_tw=max_tw,
            objective=objective,
            prefer_distribution=prefer_distribution,
        )
        week_rows = _step_state(state, demand, best_t, collect=True)
        out_rows.append(week_rows)
    return out_rows


def _final_cum_cost_from_rows(rolling_rows: list) -> float:
    if not rolling_rows:
        return 0.0
    last = rolling_rows[-1]
    return sum(last[role]["Cumulative"] for role in ROLE_NAMES)


def _demand_series_from_level(level: int) -> List[int]:
    if level == 1:
        return LOW_DEMAND_35
    if level == 3:
        return HIGH_DEMAND_35
    return BASE_DEMAND_35


def _scenario_policy_config(scenario_idx: int, selected_demand_series: Optional[List[int]] = None) -> dict:
    demand_series = selected_demand_series if selected_demand_series is not None else BASE_DEMAND_35
    if scenario_idx == 1:
        return {
            "demand_series": demand_series,
            "min_tw": 4,
            "max_tw": 8,
            "horizon": 8,
            "objective": "backlog",
            "prefer_distribution": False,
        }
    if scenario_idx == 2:
        return {
            "demand_series": demand_series,
            "min_tw": 2,
            "max_tw": 5,
            "horizon": 6,
            "objective": "total",
            "prefer_distribution": True,
        }
    return {
        "demand_series": demand_series,
        "min_tw": 1,
        "max_tw": 6,
        "horizon": 6,
        "objective": "total",
        "prefer_distribution": False,
    }


def build_beer_game_xlsx(path: str = "BeerGame_Excel_adjust.xlsx") -> None:
    wb = Workbook()
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_demand = wb.create_sheet("Demand")
    ws_summary = wb.create_sheet("Summary")

    # Inputs
    ws_inputs["A1"], ws_inputs["B1"] = "Parameter", "Value"
    ws_inputs["A3"], ws_inputs["B3"] = "InvCost", 0.25
    ws_inputs["A4"], ws_inputs["B4"] = "BackCost", 1.00
    ws_inputs["A5"], ws_inputs["B5"] = "ExpectedDemand", 100
    ws_inputs["A6"], ws_inputs["B6"] = "OrderDelay", 2
    ws_inputs["A7"], ws_inputs["B7"] = "ShipDelay", 2
    ws_inputs["A8"], ws_inputs["B8"] = "ProdDelay", 2
    ws_inputs["A10"], ws_inputs["B10"] = "Mode", "Auto"
    ws_inputs["A11"], ws_inputs["B11"] = "ScenarioIndex", 1
    ws_inputs["A12"], ws_inputs["B12"] = "DemandLevel (1=Low,2=Base,3=High)", 2
    ws_inputs["A13"], ws_inputs["B13"] = "TargetWeeks_Retailer", "=INDEX($C$27:$F$29,$B$11,1)"
    ws_inputs["A14"], ws_inputs["B14"] = "TargetWeeks_Wholesaler", "=INDEX($C$27:$F$29,$B$11,2)"
    ws_inputs["A15"], ws_inputs["B15"] = "TargetWeeks_Distributor", "=INDEX($C$27:$F$29,$B$11,3)"
    ws_inputs["A16"], ws_inputs["B16"] = "TargetWeeks_Factory", "=INDEX($C$27:$F$29,$B$11,4)"
    ws_inputs["A18"], ws_inputs["B18"] = "InitialInventory", 100
    ws_inputs["A19"], ws_inputs["B19"] = "InitialShipPipe1", 50
    ws_inputs["A20"], ws_inputs["B20"] = "InitialShipPipe2", 50
    ws_inputs["A21"], ws_inputs["B21"] = "InitialOrderPipe1", 0
    ws_inputs["A22"], ws_inputs["B22"] = "InitialOrderPipe2", 0
    ws_inputs["A23"], ws_inputs["B23"] = "DefaultOrderWeek1", 100
    ws_inputs["A24"], ws_inputs["B24"] = "ScenarioName", '=INDEX($B$27:$B$29,$B$11)'
    ws_inputs["A25"], ws_inputs["B25"] = "DemandLevelName", '=CHOOSE($B$12,"Low","Base","High")'
    selected_demand_series = _demand_series_from_level(int(ws_inputs["B12"].value))

    # Scenario presets (editable)
    ws_inputs["A26"], ws_inputs["B26"] = "ScenarioIndex", "ScenarioName"
    ws_inputs["C26"], ws_inputs["D26"], ws_inputs["E26"], ws_inputs["F26"], ws_inputs["G26"] = (
        "TargetWeeks_Retailer",
        "TargetWeeks_Wholesaler",
        "TargetWeeks_Distributor",
        "TargetWeeks_Factory",
        "Notes",
    )
    sc1_targets = _optimize_targets(
        HIGH_DEMAND_35, min_tw=4, max_tw=8, prefer_distribution=False, objective="backlog"
    )
    sc2_targets = _optimize_targets(
        BASE_DEMAND_35, min_tw=2, max_tw=5, prefer_distribution=True, objective="total"
    )
    sc3_targets = _optimize_targets(
        BASE_DEMAND_35, min_tw=1, max_tw=6, prefer_distribution=False, objective="total"
    )

    ws_inputs["A27"], ws_inputs["B27"] = 1, "Simulation - Scenario 1 - HighInv"
    ws_inputs["C27"], ws_inputs["D27"], ws_inputs["E27"], ws_inputs["F27"] = sc1_targets
    ws_inputs["G27"] = "High inventory, lowest risk of backlog"
    ws_inputs["A28"], ws_inputs["B28"] = 2, "Simulation - Scenario 2 - Balanced"
    ws_inputs["C28"], ws_inputs["D28"], ws_inputs["E28"], ws_inputs["F28"] = sc2_targets
    ws_inputs["G28"] = "Balanced inventory and backlog tradeoff"
    ws_inputs["A29"], ws_inputs["B29"] = 3, "Simulation - Scenario 3 - MinCost"
    ws_inputs["C29"], ws_inputs["D29"], ws_inputs["E29"], ws_inputs["F29"] = sc3_targets
    ws_inputs["G29"] = "Cost-focused baseline policy (editable)"

    # Demand profile (35 weeks): Base=100, Low avg=80, High avg=115
    ws_demand.append(["Week", "BaseDemand", "LowDemand", "HighDemand", "ActiveDemand"])
    for w in range(1, WEEKS + 1):
        row = w + 1
        ws_demand[f"A{row}"] = w
        ws_demand[f"B{row}"] = BASE_DEMAND_35[w - 1]
        ws_demand[f"C{row}"] = LOW_DEMAND_35[w - 1]
        ws_demand[f"D{row}"] = HIGH_DEMAND_35[w - 1]
        ws_demand[f"E{row}"] = f'=CHOOSE(Inputs!$B$12,C{row},B{row},D{row})'

    # Scenario summary for report readers
    ws_inputs["A31"] = "Scenario Summary"
    ws_inputs["A32"], ws_inputs["B32"], ws_inputs["C32"], ws_inputs["D32"], ws_inputs["E32"] = (
        "Scenario",
        "Demand Used",
        "Policy Intent",
        "Weekly KPI",
        "Final KPI",
    )
    ws_inputs["A33"], ws_inputs["B33"], ws_inputs["C33"] = (
        "Simulation - Scenario 1 - HighInv",
        "Selected in Inputs (Low/Base/High)",
        "Reduce backlog risk",
    )
    ws_inputs["A34"], ws_inputs["B34"], ws_inputs["C34"] = (
        "Simulation - Scenario 2 - Balanced",
        "Selected in Inputs (Low/Base/High)",
        "Balance inventory and backlog",
    )
    ws_inputs["A35"], ws_inputs["B35"], ws_inputs["C35"] = (
        "Simulation - Scenario 3 - MinCost",
        "Selected in Inputs (Low/Base/High)",
        "Minimize total cost",
    )
    for r in (33, 34, 35):
        ws_inputs[f"D{r}"] = "TotInvCost + TotBackCost (= TotWeekCost)"
    ws_inputs["E33"] = "=Scn1_View!AS39"
    ws_inputs["E34"] = "=Scn2_View!AS39"
    ws_inputs["E35"] = "=Scn3_View!AS39"

    headers = [
        "Week","CustomerDemand","In","Recv","BegInv","BegBacklog","TotalReq",
        "Ship","EndInv","EndBacklog","InvCost","BackCost","WeekCost","CumulativeCost",
        "OrderManual","OrderPlaced","OrderPipe1","OrderPipe2","ShipPipe1","ShipPipe2",
        "InventoryPos","TargetLevel","SuggestedOrder"
    ]

    dark = PatternFill("solid", fgColor="1F4E78")
    mid = PatternFill("solid", fgColor="D9E1F2")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Full formatted scenario views (independent calc per scenario)
    scenario_defs = [
        (1, "Simulation - Scenario 1 - HighInv"),
        (2, "Simulation - Scenario 2 - Balanced"),
        (3, "Simulation - Scenario 3 - MinCost"),
    ]

    for sc_idx, sc_title in scenario_defs:
        calc_sheets = {}
        policy_cfg = _scenario_policy_config(sc_idx, selected_demand_series=selected_demand_series)
        demand_levels = {
            1: LOW_DEMAND_35,
            2: BASE_DEMAND_35,
            3: HIGH_DEMAND_35,
        }
        for d_level, d_series in demand_levels.items():
            rolling_rows = _rolling_plan_rows(
                demand_series=d_series,
                min_tw=policy_cfg["min_tw"],
                max_tw=policy_cfg["max_tw"],
                horizon=policy_cfg["horizon"],
                objective=policy_cfg["objective"],
                prefer_distribution=policy_cfg["prefer_distribution"],
            )
            for role in ROLE_NAMES:
                calc_name = f"S{sc_idx}_D{d_level}_{role}"
                ws_calc = wb.create_sheet(calc_name)
                calc_sheets[(d_level, role)] = ws_calc
                ws_calc.append(headers)
                start = 2
                for r in range(start, start + WEEKS):
                    week = r - start + 1
                    ws_calc[f"A{r}"] = week

                    rr = rolling_rows[week - 1][role]
                    ws_calc[f"B{r}"] = rr["Incoming"]
                    ws_calc[f"C{r}"] = rr["Incoming"]
                    ws_calc[f"D{r}"] = rr["Recv"]
                    ws_calc[f"E{r}"] = rr["BegInv"]
                    ws_calc[f"F{r}"] = rr["BegBacklog"]
                    ws_calc[f"G{r}"] = rr["TotalReq"]
                    ws_calc[f"H{r}"] = rr["Ship"]
                    ws_calc[f"I{r}"] = rr["EndInv"]
                    ws_calc[f"J{r}"] = rr["EndBacklog"]
                    ws_calc[f"K{r}"] = rr["InvCost"]
                    ws_calc[f"L{r}"] = rr["BackCost"]
                    ws_calc[f"M{r}"] = rr["WeekCost"]
                    ws_calc[f"N{r}"] = rr["Cumulative"]
                    ws_calc[f"P{r}"] = rr["OrderPlaced"]
                    ws_calc[f"Q{r}"] = 0
                    ws_calc[f"R{r}"] = 0
                    ws_calc[f"S{r}"] = 0
                    ws_calc[f"T{r}"] = 0
                    ws_calc[f"U{r}"] = 0
                    ws_calc[f"V{r}"] = 0
                    ws_calc[f"W{r}"] = 0

                ws_calc.sheet_state = "hidden"

        ws_sc = wb.create_sheet(f"Scn{sc_idx}_View")
        ws_sc["A1"] = sc_title
        ws_sc["A2"] = "Edit Inputs to change demand, costs, and policy parameters."
        ws_sc["A4"] = "Week"

        role_metric_headers = [
            "Incoming order", "Incoming shipment", "Outgoing shipment",
            "Inventory", "Backlog", "Order placed",
            "InvCost", "BackCost", "WeekCost", "Cumulative"
        ]
        role_metric_map = {
            "Incoming order": "C",
            "Incoming shipment": "D",
            "Outgoing shipment": "H",
            "Inventory": "I",
            "Backlog": "J",
            "Order placed": "P",
            "InvCost": "K",
            "BackCost": "L",
            "WeekCost": "M",
            "Cumulative": "N"
        }
        block_starts = {"Retailer": 2, "Wholesaler": 12, "Distributor": 22, "Factory": 32}

        ws_sc.merge_cells("A3:A4")
        for role in ROLE_NAMES:
            c_start_idx = block_starts[role]
            c_end_idx = c_start_idx + len(role_metric_headers) - 1
            c_start = get_column_letter(c_start_idx)
            c_end = get_column_letter(c_end_idx)
            ws_sc.merge_cells(f"{c_start}3:{c_end}3")
            ws_sc[f"{c_start}3"] = role
            for i, h in enumerate(role_metric_headers):
                ws_sc[f"{get_column_letter(c_start_idx + i)}4"] = h

        ws_sc.merge_cells("AP3:AS3")
        ws_sc["AP3"] = "Total"
        ws_sc["AP4"] = "TotInvCost"
        ws_sc["AQ4"] = "TotBackCost"
        ws_sc["AR4"] = "TotWeekCost"
        ws_sc["AS4"] = "CumCost"

        for w in range(1, WEEKS + 1):
            vr = 4 + w
            rr = 1 + w
            ws_sc[f"A{vr}"] = w
            for role in ROLE_NAMES:
                c_start_idx = block_starts[role]
                for i, h in enumerate(role_metric_headers):
                    mcol = role_metric_map[h]
                    ws_sc[f"{get_column_letter(c_start_idx + i)}{vr}"] = (
                        f"=CHOOSE(Inputs!$B$12,"
                        f"'S{sc_idx}_D1_{role}'!{mcol}{rr},"
                        f"'S{sc_idx}_D2_{role}'!{mcol}{rr},"
                        f"'S{sc_idx}_D3_{role}'!{mcol}{rr})"
                    )

            ws_sc[f"AP{vr}"] = f"=SUM(H{vr},R{vr},AB{vr},AL{vr})"
            ws_sc[f"AQ{vr}"] = f"=SUM(I{vr},S{vr},AC{vr},AM{vr})"
            ws_sc[f"AR{vr}"] = f"=SUM(J{vr},T{vr},AD{vr},AN{vr})"
            ws_sc[f"AS{vr}"] = f"=AR{vr}" if w == 1 else f"=AS{vr-1}+AR{vr}"

        ws_sc["A1"].font = Font(bold=True, size=14)
        ws_sc["A2"].font = Font(italic=True, color="666666")
        for c in ["A3", "B3", "L3", "V3", "AF3", "AP3"]:
            ws_sc[c].fill = dark
            ws_sc[c].font = white_font
            ws_sc[c].alignment = center

        for row in ws_sc.iter_rows(min_row=4, max_row=4 + WEEKS, min_col=1, max_col=45):
            for c in row:
                c.border = border
                c.alignment = center
        for c in ws_sc[4]:
            c.fill = mid
            c.font = bold

        ws_sc.freeze_panes = "B5"
        ws_sc.column_dimensions["A"].width = 8
        for col_idx in range(2, 46):
            ws_sc.column_dimensions[get_column_letter(col_idx)].width = 10

    # Dedicated summary sheet (matrix layout)
    ws_summary["A1"] = "Demand Used"
    ws_summary["B1"] = "High demand profile"
    ws_summary["C1"] = "Base demand profile"
    ws_summary["D1"] = "Low demand profile"
    ws_summary["A2"] = "Scenario"
    ws_summary["A3"] = "Simulation - Scenario 1 - HighInv"
    ws_summary["A4"] = "Simulation - Scenario 2 - Balanced"
    ws_summary["A5"] = "Simulation - Scenario 3 - MinCost"

    # Matrix values: final KPI (system cumulative cost at week 35)
    demand_profiles = [HIGH_DEMAND_35, BASE_DEMAND_35, LOW_DEMAND_35]
    summary_rows = {1: 3, 2: 4, 3: 5}
    summary_cols = {0: "B", 1: "C", 2: "D"}
    for sc_idx in (1, 2, 3):
        cfg = _scenario_policy_config(sc_idx, selected_demand_series=selected_demand_series)
        for d_idx, demand_series in enumerate(demand_profiles):
            rows = _rolling_plan_rows(
                demand_series=demand_series,
                min_tw=cfg["min_tw"],
                max_tw=cfg["max_tw"],
                horizon=cfg["horizon"],
                objective=cfg["objective"],
                prefer_distribution=cfg["prefer_distribution"],
            )
            ws_summary[f"{summary_cols[d_idx]}{summary_rows[sc_idx]}"] = round(
                _final_cum_cost_from_rows(rows), 2
            )

    for c in ws_summary[1]:
        c.fill = mid
        c.font = bold
        c.alignment = center
        c.border = border
    for c in ws_summary[2]:
        c.fill = mid
        c.font = bold
        c.alignment = center
        c.border = border
    for row in ws_summary.iter_rows(min_row=3, max_row=5, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = center
            cell.border = border
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 28
    ws_summary.column_dimensions["C"].width = 28
    ws_summary.column_dimensions["D"].width = 28

    wb.save(path)

if __name__ == "__main__":
    build_beer_game_xlsx()
